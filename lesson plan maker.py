import openpyxl

#constants 
subject  = 'Listening'
ESL_level =  str(300)
write_to = "lesson_plans_Fall2018//listening 300"
spreadsheet_file = "listening 300 Schdule spreadsheet.xlsx"


wb = openpyxl.load_workbook(spreadsheet_file)
sheet = wb.active
'''for row in range(1, sheet.max_row):
	print(list(sheet.columns)[0][row].value)'''

for i in range (1,22):
	week = str(i)
	page_number     =  str((list(sheet.columns)[0][i].value))
	information     =  str((list(sheet.columns)[1][i].value))
	verification    =  str((list(sheet.columns)[2][i].value))
	activities      =  str((list(sheet.columns)[3][i].value))
	worksheets      =  str((list(sheet.columns)[4][i].value))
	materials       =  str((list(sheet.columns)[5][i].value))
	additional      =  str((list(sheet.columns)[6][i].value))
	other_resources =  str((list(sheet.columns)[7][i].value))

	lesson_plan = open('{0}//lesson plan {1} plan week {2}.txt'.format(write_to, subject, week), 'w')

	lesson_plan.write('Lizen High School – Foreign Language Center'.center(75, ' '))
	lesson_plan.write('\n' + ('Lesson Plan for Week ' + week).center(75,' '))
	lesson_plan.write('\n'*2 + 'ESL level: ' + ESL_level + ' '*20 + 'Subject: ' + subject + ' '*20 + 'Prepared by: Daniel')
	lesson_plan.write('\n'*3 + 'Story Lesson (page): ' + page_number + '\n Information: ' + information +' \n Verification: ' + verification +' \n Activities: ' + activities + ' \n Worksheets: ' + worksheets + '\n Materials needed: ' + materials + ' \n Additional Notes: ' + additional + '\n Other Resources: ' + other_resources + '\n' + '-'*100 + '\n' + 'Summmary: '.center(75, ' '))
	#print('This is a {0} test {1}'.format(subject, week))

'''wb = openpyxl.load_workbook('credits_spreadsheet.xlsx')
	sheet = wb.active
	for credit in range (1,sheet.max_row):
		for row in range (3):
			print(list(sheet.rows)[0][row].value +': ' + list(sheet.rows)[credit][row].value)
		'''